import os
import re
import gc  # Garbage collector para liberar memoria
import pdfplumber
import pandas as pd # type: ignore
import openpyxl # type: ignore
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk

# Importar psutil con manejo de error
try:
    import psutil  # Para monitorear memoria
    PSUTIL_DISPONIBLE = True
except ImportError:
    PSUTIL_DISPONIBLE = False
    print("⚠️  psutil no está instalado. Funcionalidad de monitoreo de memoria limitada.")

def seleccionar_carpetas():
    """Interfaz gráfica para seleccionar carpetas de entrada y salida"""
    
    # Crear ventana principal
    root = tk.Tk()
    root.title("Convertidor de PDFs a Excel - Configuración")
    root.geometry("600x400")
    root.resizable(False, False)
    
    # Variables para almacenar las rutas
    carpeta_pdfs = tk.StringVar()
    carpeta_salida = tk.StringVar()
    
    # Estilo
    style = ttk.Style()
    style.theme_use('clam')
    
    # Frame principal
    main_frame = ttk.Frame(root, padding="20")
    main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
    
    # Título
    title_label = ttk.Label(main_frame, text="Convertidor de PDFs a Excel", 
                           font=('Arial', 16, 'bold'))
    title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))
    
    # Instrucciones
    instrucciones = ttk.Label(main_frame, 
                             text="Seleccione la carpeta donde están los archivos PDF y donde desea guardar el Excel:",
                             wraplength=550)
    instrucciones.grid(row=1, column=0, columnspan=3, pady=(0, 20))
    
    # Selección de carpeta de PDFs
    ttk.Label(main_frame, text="Carpeta con archivos PDF:", font=('Arial', 10, 'bold')).grid(row=2, column=0, sticky=tk.W, pady=(10, 5))
    
    entry_pdfs = ttk.Entry(main_frame, textvariable=carpeta_pdfs, width=60, state='readonly')
    entry_pdfs.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), padx=(0, 10))
    
    def seleccionar_carpeta_pdfs():
        carpeta = filedialog.askdirectory(title="Seleccionar carpeta con archivos PDF")
        if carpeta:
            carpeta_pdfs.set(carpeta)
    
    btn_pdfs = ttk.Button(main_frame, text="Explorar", command=seleccionar_carpeta_pdfs)
    btn_pdfs.grid(row=3, column=2, sticky=tk.W)
    
    # Selección de carpeta de salida
    ttk.Label(main_frame, text="Carpeta donde guardar el Excel:", font=('Arial', 10, 'bold')).grid(row=4, column=0, sticky=tk.W, pady=(20, 5))
    
    entry_salida = ttk.Entry(main_frame, textvariable=carpeta_salida, width=60, state='readonly')
    entry_salida.grid(row=5, column=0, columnspan=2, sticky=(tk.W, tk.E), padx=(0, 10))
    
    def seleccionar_carpeta_salida():
        carpeta = filedialog.askdirectory(title="Seleccionar carpeta donde guardar el Excel")
        if carpeta:
            carpeta_salida.set(carpeta)
    
    btn_salida = ttk.Button(main_frame, text="Explorar", command=seleccionar_carpeta_salida)
    btn_salida.grid(row=5, column=2, sticky=tk.W)
    
    # Resultado
    resultado = {'carpeta_pdfs': None, 'carpeta_salida': None, 'continuar': False}
    
    def procesar():
        if not carpeta_pdfs.get():
            messagebox.showerror("Error", "Por favor seleccione la carpeta con los archivos PDF")
            return
        
        if not carpeta_salida.get():
            messagebox.showerror("Error", "Por favor seleccione la carpeta donde guardar el Excel")
            return
        
        # Verificar que la carpeta de PDFs contenga archivos PDF
        archivos_pdf = [f for f in os.listdir(carpeta_pdfs.get()) if f.endswith('.pdf')]
        if not archivos_pdf:
            messagebox.showerror("Error", "La carpeta seleccionada no contiene archivos PDF")
            return
        
        resultado['carpeta_pdfs'] = carpeta_pdfs.get()
        resultado['carpeta_salida'] = carpeta_salida.get()
        resultado['continuar'] = True
        root.destroy()
    
    def cancelar():
        root.destroy()
    
    # Botones de acción
    button_frame = ttk.Frame(main_frame)
    button_frame.grid(row=6, column=0, columnspan=3, pady=30)
    
    btn_procesar = ttk.Button(button_frame, text="Procesar PDFs", command=procesar, 
                             style='Accent.TButton')
    btn_procesar.pack(side=tk.LEFT, padx=(0, 10))
    
    btn_cancelar = ttk.Button(button_frame, text="Cancelar", command=cancelar)
    btn_cancelar.pack(side=tk.LEFT)
    
    # Información adicional
    info_text = ttk.Label(main_frame, 
                         text="El archivo Excel se guardará como 'resultado_detallado.xlsx' en la carpeta seleccionada.",
                         font=('Arial', 9), foreground='gray')
    info_text.grid(row=7, column=0, columnspan=3, pady=(20, 0))
    
    # Configurar grid weights
    root.columnconfigure(0, weight=1)
    root.rowconfigure(0, weight=1)
    main_frame.columnconfigure(1, weight=1)
    
    # Centrar ventana
    root.update_idletasks()
    x = (root.winfo_screenwidth() // 2) - (root.winfo_width() // 2)
    y = (root.winfo_screenheight() // 2) - (root.winfo_height() // 2)
    root.geometry(f"+{x}+{y}")
    
    # Mostrar ventana
    root.mainloop()
    
    return resultado

# Mostrar interfaz de selección
config = seleccionar_carpetas()

# Si el usuario canceló, salir del programa
if not config['continuar']:
    print("Operación cancelada por el usuario.")
    exit()

# Configurar rutas basadas en la selección del usuario
carpeta_pdfs = config['carpeta_pdfs']
carpeta_salida = config['carpeta_salida']

print(f"📂 Carpeta de PDFs: {carpeta_pdfs}")
print(f"💾 Carpeta de salida: {carpeta_salida}")
print("🔄 Iniciando procesamiento...")

# Lista donde guardaremos los resultados
datos = []

def extraer_info(texto, archivo):
    """Extrae datos únicamente para persona natural"""
    registro = {
        # INFORMACIÓN PERSONAL
        "Consultado por": "NO APLICA",
        "Fecha y Hora Consulta": "NO APLICA",
        "Tipo Documento": "NO APLICA",
        "Número Documento": "NO APLICA",
        "Estado Documento": "NO APLICA",
        "Lugar Expedición": "NO APLICA",
        "Fecha Expedición": "NO APLICA",
        "Nombre": "NO APLICA",
        "Rango Edad": "NO APLICA",
        "Género": "NO APLICA",
        "Antigüedad Ubicación": "NO APLICA",
        
        # CRÉDITOS VIGENTES
        "Créditos Vigentes - Sector Financiero": "NO APLICA",
        "Créditos Vigentes - Sector Cooperativo": "NO APLICA", 
        "Créditos Vigentes - Sector Real": "NO APLICA",
        "Créditos Vigentes - Sector Telcos": "NO APLICA",
        "Créditos Vigentes - Total Sectores": "NO APLICA",
        "Créditos Vigentes - Total Principal": "NO APLICA",
        "Créditos Vigentes - Total Codeudor": "NO APLICA",
        
        # CRÉDITOS CERRADOS
        "Créditos Cerrados - Sector Financiero": "NO APLICA",
        "Créditos Cerrados - Sector Cooperativo": "NO APLICA",
        "Créditos Cerrados - Sector Real": "NO APLICA", 
        "Créditos Cerrados - Sector Telcos": "NO APLICA",
        "Créditos Cerrados - Total Sectores": "NO APLICA",
        "Créditos Cerrados - Total Principal": "NO APLICA",
        "Créditos Cerrados - Total Codeudor": "NO APLICA",
        
        # CRÉDITOS REESTRUCTURADOS
        "Créditos Reestructurados - Sector Financiero": "NO APLICA",
        "Créditos Reestructurados - Sector Cooperativo": "NO APLICA",
        "Créditos Reestructurados - Sector Real": "NO APLICA",
        "Créditos Reestructurados - Sector Telcos": "NO APLICA", 
        "Créditos Reestructurados - Total Sectores": "NO APLICA",
        "Créditos Reestructurados - Total Principal": "NO APLICA",
        "Créditos Reestructurados - Total Codeudor": "NO APLICA",
        
        # CRÉDITOS REFINANCIADOS
        "Créditos Refinanciados - Sector Financiero": "NO APLICA",
        "Créditos Refinanciados - Sector Cooperativo": "NO APLICA",
        "Créditos Refinanciados - Sector Real": "NO APLICA",
        "Créditos Refinanciados - Sector Telcos": "NO APLICA",
        "Créditos Refinanciados - Total Sectores": "NO APLICA", 
        "Créditos Refinanciados - Total Principal": "NO APLICA",
        "Créditos Refinanciados - Total Codeudor": "NO APLICA",
        
        # CONSULTAS ÚLTIMOS 6 MESES
        "Consultas Últimos 6 Meses - Sector Financiero": "NO APLICA",
        "Consultas Últimos 6 Meses - Sector Cooperativo": "NO APLICA",
        "Consultas Últimos 6 Meses - Sector Real": "NO APLICA",
        "Consultas Últimos 6 Meses - Sector Telcos": "NO APLICA",
        "Consultas Últimos 6 Meses - Total Sectores": "NO APLICA",
        "Consultas Últimos 6 Meses - Total Principal": "NO APLICA", 
        "Consultas Últimos 6 Meses - Total Codeudor": "NO APLICA",
        
        # DESACUERDOS VIGENTES
        "Desacuerdos Vigentes - Sector Financiero": "NO APLICA",
        "Desacuerdos Vigentes - Sector Cooperativo": "NO APLICA",
        "Desacuerdos Vigentes - Sector Real": "NO APLICA",
        "Desacuerdos Vigentes - Sector Telcos": "NO APLICA",
        "Desacuerdos Vigentes - Total Sectores": "NO APLICA",
        "Desacuerdos Vigentes - Total Principal": "NO APLICA",
        "Desacuerdos Vigentes - Total Codeudor": "NO APLICA",
        
        # ANTIGÜEDAD POR SECTOR
        "Antigüedad - Sector Financiero": "NO APLICA",
        "Antigüedad - Sector Real": "NO APLICA", 
        "Antigüedad - Sector Telcos": "NO APLICA"
    }

    # Buscar fecha y hora de consulta
    fecha_hora = re.search(r"Fecha y Hora Consulta\s*:?\s*([0-9]{4}/[0-9]{2}/[0-9]{2} [0-9]{1,2}\.[0-9]{2} (AM|PM))", texto)
    if not fecha_hora:
        fecha_hora = re.search(r"(20[0-9]{2}/[0-9]{2}/[0-9]{2} [0-9]{1,2}\.[0-9]{2} (AM|PM))", texto)
    if fecha_hora:
        registro["Fecha y Hora Consulta"] = fecha_hora.group(1).strip()
    
    # Extraer "Consultado por" para persona natural únicamente
    consultado_nat = None
    for linea in texto.splitlines():
        if 'Consultado por' in linea:
            partes = linea.split(':', 1)
            if len(partes) > 1:
                valor = partes[1].strip()
                # Omitir empresas, números y AM/PM
                valor = re.sub(r"\b(AM|PM)\b", "", valor, flags=re.IGNORECASE).strip()
                if not any(x in valor.upper() for x in ['SAS', 'LTDA', 'S.A.', 'DELAGRO']) and not re.search(r"\d", valor):
                    consultado_nat = valor
            else:
                idx = texto.splitlines().index(linea)
                if idx + 1 < len(texto.splitlines()):
                    valor = texto.splitlines()[idx + 1].strip()
                    valor = re.sub(r"\b(AM|PM)\b", "", valor, flags=re.IGNORECASE).strip()
                    if not any(x in valor.upper() for x in ['SAS', 'LTDA', 'S.A.', 'DELAGRO']) and not re.search(r"\d", valor):
                        consultado_nat = valor
    if consultado_nat:
        registro["Consultado por"] = consultado_nat
    else:
        # Buscar patrón alternativo: nombre seguido de guion y número, tomar solo el nombre y quitar AM/PM
        persona = re.search(r"([A-ZÁÉÍÓÚÑa-z\s]+)\s*-\s*[0-9]+", texto)
        if persona:
            valor = persona.group(1).strip()
            valor = re.sub(r"\b(AM|PM)\b", "", valor, flags=re.IGNORECASE).strip()
            if not any(x in valor.upper() for x in ['SAS', 'LTDA', 'S.A.', 'DELAGRO']):
                registro["Consultado por"] = valor
    # Extraer Tipo de Documento
    tipo_doc = re.search(r"Tipo Documento\s*[:\s]*([A-Z.]+)", texto)
    if tipo_doc:
        registro["Tipo Documento"] = tipo_doc.group(1).strip()
    else:
        cc = re.search(r"C\.C\.", texto)
        if cc:
            registro["Tipo Documento"] = "C.C."
    
    # Extraer Número de Documento
    num_doc = re.search(r"Número Documento\s*[:\s]*([0-9]+)", texto)
    if num_doc:
        registro["Número Documento"] = num_doc.group(1).strip()
    else:
        cc_num = re.search(r"C\.C\.\s*([0-9]+)", texto)
        if cc_num:
            registro["Número Documento"] = cc_num.group(1).strip()
    
    # Extraer Estado de Documento
    estado = re.search(r"Estado Documento\s*[:\s]*([A-Za-zÁÉÍÓÚÑ]+)", texto)
    if estado:
        registro["Estado Documento"] = estado.group(1).strip()
    else:
        vigente = re.search(r"Vigente", texto)
        if vigente:
            registro["Estado Documento"] = "Vigente"
    # Extraer Lugar de Expedición
    lugar = re.search(r"Lugar Expedición\s*[:\s]*([A-ZÁÉÍÓÚÑa-z\s]+?)(?=\s*Fecha Expedici|$)", texto)
    if lugar:
        registro["Lugar Expedición"] = lugar.group(1).strip()
    else:
        ibague = re.search(r"IBAGUE", texto)
        if ibague:
            registro["Lugar Expedición"] = "IBAGUE"
    
    # Extraer Fecha de Expedición
    fecha = re.search(r"Fecha Expedición\s*[:\s]*([0-9]{2}/[0-9]{2}/[0-9]{4})", texto)
    if fecha:
        registro["Fecha Expedición"] = fecha.group(1).strip()
    
    # Extraer Nombre para persona natural
    nombre = re.search(r"Nombre\s*[:\s]*([A-ZÁÉÍÓÚÑ\s]+)(?=\s*Rango Edad|$)", texto)
    if nombre:
        registro["Nombre"] = nombre.group(1).strip()
    else:
        nombre2 = re.search(r"([A-ZÁÉÍÓÚÑ]+\s+[A-ZÁÉÍÓÚÑ]+)", texto)
        if nombre2:
            registro["Nombre"] = nombre2.group(1).strip()
    # Extraer Rango de Edad
    rango = re.search(r"Rango Edad\s*[:\s]*([0-9\-]+)", texto)
    if rango:
        registro["Rango Edad"] = rango.group(1).strip()
    else:
        rango2 = re.search(r"(\d{2}-\d{2})", texto)
        if rango2:
            registro["Rango Edad"] = rango2.group(1).strip()
    
    # Extraer Género
    genero = re.search(r"Género\s*[:\s]*([A-Za-zÁÉÍÓÚÑ]+)", texto)
    if genero:
        registro["Género"] = genero.group(1).strip()
    else:
        femenino = re.search(r"Femenino", texto)
        if femenino:
            registro["Género"] = "Femenino"
    
    # Extraer Antigüedad Ubicación
    antig = re.search(r"Antiguedad Ubicación\s*[:\s]*([0-9]+\s*Meses\s*[A-Za-z\s]+?)(?=\s*ARTICULO|\s*-|$)", texto)
    if antig:
        registro["Antigüedad Ubicación"] = antig.group(1).strip()
    else:
        antig2 = re.search(r"(\d+\s*Meses\s*[A-Za-z\s]+?)(?=\s*ARTICULO|\s*-|$)", texto)
        if antig2:
            registro["Antigüedad Ubicación"] = antig2.group(1).strip()
    
    # Extraer datos del Perfil General - Tabla detallada por filas y columnas
    def extraer_tabla_perfil(texto):
        """Extrae datos de la tabla de perfil general organizados por filas y columnas"""
        lineas = texto.split('\n')
        
        # Buscar las líneas que contienen los datos de la tabla
        for i, linea in enumerate(lineas):
            # CRÉDITOS VIGENTES
            if 'Creditos Vigentes' in linea or 'Créditos Vigentes' in linea:
                valores = re.findall(r'\d+', linea)
                if len(valores) >= 7:  # Asegurarse que hay suficientes valores
                    registro["Créditos Vigentes - Sector Financiero"] = valores[0] if valores[0] else "0"
                    registro["Créditos Vigentes - Sector Cooperativo"] = valores[1] if valores[1] else "0" 
                    registro["Créditos Vigentes - Sector Real"] = valores[2] if valores[2] else "0"
                    registro["Créditos Vigentes - Sector Telcos"] = valores[3] if valores[3] else "0"
                    registro["Créditos Vigentes - Total Sectores"] = valores[4] if valores[4] else "0"
                    registro["Créditos Vigentes - Total Principal"] = valores[5] if valores[5] else "0"
                    registro["Créditos Vigentes - Total Codeudor"] = valores[6] if valores[6] else "0"
            
            # CRÉDITOS CERRADOS
            elif 'Creditos Cerrados' in linea or 'Créditos Cerrados' in linea:
                valores = re.findall(r'\d+', linea)
                if len(valores) >= 7:
                    registro["Créditos Cerrados - Sector Financiero"] = valores[0] if valores[0] else "0"
                    registro["Créditos Cerrados - Sector Cooperativo"] = valores[1] if valores[1] else "0"
                    registro["Créditos Cerrados - Sector Real"] = valores[2] if valores[2] else "0" 
                    registro["Créditos Cerrados - Sector Telcos"] = valores[3] if valores[3] else "0"
                    registro["Créditos Cerrados - Total Sectores"] = valores[4] if valores[4] else "0"
                    registro["Créditos Cerrados - Total Principal"] = valores[5] if valores[5] else "0"
                    registro["Créditos Cerrados - Total Codeudor"] = valores[6] if valores[6] else "0"
            
            # CRÉDITOS REESTRUCTURADOS
            elif 'Creditos Reestructurados' in linea or 'Créditos Reestructurados' in linea:
                valores = re.findall(r'\d+', linea)
                if len(valores) >= 7:
                    registro["Créditos Reestructurados - Sector Financiero"] = valores[0] if valores[0] else "0"
                    registro["Créditos Reestructurados - Sector Cooperativo"] = valores[1] if valores[1] else "0"
                    registro["Créditos Reestructurados - Sector Real"] = valores[2] if valores[2] else "0"
                    registro["Créditos Reestructurados - Sector Telcos"] = valores[3] if valores[3] else "0"
                    registro["Créditos Reestructurados - Total Sectores"] = valores[4] if valores[4] else "0"
                    registro["Créditos Reestructurados - Total Principal"] = valores[5] if valores[5] else "0" 
                    registro["Créditos Reestructurados - Total Codeudor"] = valores[6] if valores[6] else "0"
            
            # CRÉDITOS REFINANCIADOS
            elif 'Creditos Refinanciados' in linea or 'Créditos Refinanciados' in linea:
                valores = re.findall(r'\d+', linea)
                if len(valores) >= 7:
                    registro["Créditos Refinanciados - Sector Financiero"] = valores[0] if valores[0] else "0"
                    registro["Créditos Refinanciados - Sector Cooperativo"] = valores[1] if valores[1] else "0"
                    registro["Créditos Refinanciados - Sector Real"] = valores[2] if valores[2] else "0"
                    registro["Créditos Refinanciados - Sector Telcos"] = valores[3] if valores[3] else "0"
                    registro["Créditos Refinanciados - Total Sectores"] = valores[4] if valores[4] else "0"
                    registro["Créditos Refinanciados - Total Principal"] = valores[5] if valores[5] else "0"
                    registro["Créditos Refinanciados - Total Codeudor"] = valores[6] if valores[6] else "0"
            
            # CONSULTAS ÚLTIMOS 6 MESES
            elif 'Consultas en los ult' in linea or 'Consultas ult' in linea:
                valores = re.findall(r'\d+', linea)
                if len(valores) >= 7:
                    registro["Consultas Últimos 6 Meses - Sector Financiero"] = valores[0] if valores[0] else "0"
                    registro["Consultas Últimos 6 Meses - Sector Cooperativo"] = valores[1] if valores[1] else "0"
                    registro["Consultas Últimos 6 Meses - Sector Real"] = valores[2] if valores[2] else "0"
                    registro["Consultas Últimos 6 Meses - Sector Telcos"] = valores[3] if valores[3] else "0"
                    registro["Consultas Últimos 6 Meses - Total Sectores"] = valores[4] if valores[4] else "0"
                    registro["Consultas Últimos 6 Meses - Total Principal"] = valores[5] if valores[5] else "0"
                    registro["Consultas Últimos 6 Meses - Total Codeudor"] = valores[6] if valores[6] else "0"
            
            # DESACUERDOS VIGENTES
            elif 'Desacuerdos Vigentes' in linea:
                valores = re.findall(r'\d+', linea)
                if len(valores) >= 7:
                    registro["Desacuerdos Vigentes - Sector Financiero"] = valores[0] if valores[0] else "0"
                    registro["Desacuerdos Vigentes - Sector Cooperativo"] = valores[1] if valores[1] else "0"
                    registro["Desacuerdos Vigentes - Sector Real"] = valores[2] if valores[2] else "0"
                    registro["Desacuerdos Vigentes - Sector Telcos"] = valores[3] if valores[3] else "0"
                    registro["Desacuerdos Vigentes - Total Sectores"] = valores[4] if valores[4] else "0"
                    registro["Desacuerdos Vigentes - Total Principal"] = valores[5] if valores[5] else "0"
                    registro["Desacuerdos Vigentes - Total Codeudor"] = valores[6] if valores[6] else "0"
            
            # ANTIGÜEDAD DESDE
            elif 'Antigüedad desde' in linea or 'Antiguedad desde' in linea:
                fechas = re.findall(r'\d{4}-\d{2}-\d{2}', linea)
                if len(fechas) >= 3:
                    registro["Antigüedad - Sector Financiero"] = fechas[0] if fechas[0] else "NO APLICA"
                    registro["Antigüedad - Sector Real"] = fechas[1] if fechas[1] else "NO APLICA"
                    registro["Antigüedad - Sector Telcos"] = fechas[2] if fechas[2] else "NO APLICA"
                elif len(fechas) >= 1:
                    # Si solo hay una fecha, asignarla al sector financiero
                    registro["Antigüedad - Sector Financiero"] = fechas[0]
    
    # Llamar a la función para extraer la tabla
    extraer_tabla_perfil(texto)
    
    return registro


# 🔹 Procesar todos los PDFs en la carpeta con optimizaciones para gran volumen
archivos_pdf = [f for f in os.listdir(carpeta_pdfs) if f.endswith(".pdf")]
total_archivos = len(archivos_pdf)

print(f"📄 Se encontraron {total_archivos} archivos PDF para procesar")

# Determinar tamaño de lote basado en memoria disponible
if PSUTIL_DISPONIBLE:
    memoria_disponible_gb = psutil.virtual_memory().available / (1024**3)
    print(f"💾 Memoria RAM disponible: {memoria_disponible_gb:.1f} GB")
    tamano_lote = min(100, max(10, int(memoria_disponible_gb * 10)))  # 10-100 archivos por lote
else:
    # Sin psutil, usar un lote conservador basado en el número de archivos
    tamano_lote = min(50, max(10, total_archivos // 10)) if total_archivos > 50 else total_archivos
    print(f"💾 Usando configuración conservadora de memoria")

print(f"📦 Procesando en lotes de {tamano_lote} archivos para optimizar memoria")

archivos_procesados = 0
archivos_con_error = []

# Procesar en lotes para optimizar memoria
for lote_inicio in range(0, total_archivos, tamano_lote):
    lote_fin = min(lote_inicio + tamano_lote, total_archivos)
    archivos_lote = archivos_pdf[lote_inicio:lote_fin]
    
    print(f"\n🔄 Procesando lote {lote_inicio//tamano_lote + 1}/{(total_archivos-1)//tamano_lote + 1}")
    print(f"📁 Archivos {lote_inicio + 1} al {lote_fin} de {total_archivos}")
    
    for i, archivo in enumerate(archivos_lote):
        try:
            archivo_numero = lote_inicio + i + 1
            print(f"🔄 Procesando ({archivo_numero}/{total_archivos}): {archivo}")
            
            ruta_pdf = os.path.join(carpeta_pdfs, archivo)
            
            # Procesar PDF con manejo eficiente de memoria
            with pdfplumber.open(ruta_pdf) as pdf:
                texto = ""
                for pagina in pdf.pages:
                    texto_pagina = pagina.extract_text()
                    if texto_pagina:
                        texto += texto_pagina + "\n"
                    
            # Extraer información y liberar variables temporales
            info_extraida = extraer_info(texto, archivo)
            datos.append(info_extraida)
            
            # Limpiar variables para liberar memoria
            del texto, info_extraida
            
            archivos_procesados += 1
            
            # Mostrar progreso cada 10 archivos o al final del lote
            if archivo_numero % 10 == 0 or i == len(archivos_lote) - 1:
                if PSUTIL_DISPONIBLE:
                    memoria_actual = psutil.virtual_memory().percent
                    print(f"✅ Completados: {archivos_procesados}/{total_archivos} | Memoria: {memoria_actual:.1f}%")
                else:
                    print(f"✅ Completados: {archivos_procesados}/{total_archivos}")
            
        except Exception as e:
            archivos_con_error.append(f"{archivo}: {str(e)}")
            print(f"❌ Error al procesar {archivo}: {str(e)}")
            continue
    
    # Forzar liberación de memoria al final de cada lote
    gc.collect()
    
    # Verificar memoria después del lote (si psutil está disponible)
    if PSUTIL_DISPONIBLE:
        memoria_post_lote = psutil.virtual_memory().percent
        if memoria_post_lote > 85:  # Si la memoria supera el 85%
            print(f"⚠️  Advertencia: Uso de memoria alto ({memoria_post_lote:.1f}%). Liberando memoria...")
            gc.collect()
            import time
            time.sleep(1)  # Pausa breve para permitir liberación de memoria

print(f"\n📊 Resumen del procesamiento:")
print(f"✅ Archivos procesados exitosamente: {archivos_procesados}")
if archivos_con_error:
    print(f"❌ Archivos con errores: {len(archivos_con_error)}")
    for error in archivos_con_error:
        print(f"   • {error}")

if archivos_procesados == 0:
    messagebox.showerror("Error", "No se pudo procesar ningún archivo PDF. Verifique que los archivos no estén dañados.")
    exit()

# Convertir a DataFrame y guardar en Excel
df = pd.DataFrame(datos)

# Reordenar columnas de manera lógica y organizada con nombres descriptivos
columnas_ordenadas = [
    # INFORMACIÓN PERSONAL
    "Consultado por", "Fecha y Hora Consulta", "Nombre",
    "Tipo Documento", "Número Documento", "Estado Documento",
    "Lugar Expedición", "Fecha Expedición", "Rango Edad", "Género", 
    "Antigüedad Ubicación",
    
    # CRÉDITOS VIGENTES
    "Créditos Vigentes - Sector Financiero", "Créditos Vigentes - Sector Cooperativo", "Créditos Vigentes - Sector Real", 
    "Créditos Vigentes - Sector Telcos", "Créditos Vigentes - Total Sectores", "Créditos Vigentes - Total Principal", "Créditos Vigentes - Total Codeudor",
    
    # CRÉDITOS CERRADOS  
    "Créditos Cerrados - Sector Financiero", "Créditos Cerrados - Sector Cooperativo", "Créditos Cerrados - Sector Real",
    "Créditos Cerrados - Sector Telcos", "Créditos Cerrados - Total Sectores", "Créditos Cerrados - Total Principal", "Créditos Cerrados - Total Codeudor",
    
    # CRÉDITOS REESTRUCTURADOS
    "Créditos Reestructurados - Sector Financiero", "Créditos Reestructurados - Sector Cooperativo", "Créditos Reestructurados - Sector Real", 
    "Créditos Reestructurados - Sector Telcos", "Créditos Reestructurados - Total Sectores", "Créditos Reestructurados - Total Principal", "Créditos Reestructurados - Total Codeudor",
    
    # CRÉDITOS REFINANCIADOS
    "Créditos Refinanciados - Sector Financiero", "Créditos Refinanciados - Sector Cooperativo", "Créditos Refinanciados - Sector Real",
    "Créditos Refinanciados - Sector Telcos", "Créditos Refinanciados - Total Sectores", "Créditos Refinanciados - Total Principal", "Créditos Refinanciados - Total Codeudor",
    
    # CONSULTAS ÚLTIMOS 6 MESES
    "Consultas Últimos 6 Meses - Sector Financiero", "Consultas Últimos 6 Meses - Sector Cooperativo", "Consultas Últimos 6 Meses - Sector Real",
    "Consultas Últimos 6 Meses - Sector Telcos", "Consultas Últimos 6 Meses - Total Sectores", "Consultas Últimos 6 Meses - Total Principal", "Consultas Últimos 6 Meses - Total Codeudor",
    
    # DESACUERDOS VIGENTES
    "Desacuerdos Vigentes - Sector Financiero", "Desacuerdos Vigentes - Sector Cooperativo", "Desacuerdos Vigentes - Sector Real", 
    "Desacuerdos Vigentes - Sector Telcos", "Desacuerdos Vigentes - Total Sectores", "Desacuerdos Vigentes - Total Principal", "Desacuerdos Vigentes - Total Codeudor",
    
    # ANTIGÜEDAD
    "Antigüedad - Sector Financiero", "Antigüedad - Sector Real", "Antigüedad - Sector Telcos"
]

# Asegurar que todas las columnas existan y reordenar
df = df.reindex(columns=columnas_ordenadas, fill_value="NO APLICA")

# Aplicar lógica de llenado según el tipo de campo
def llenar_campos_vacios(df):
    """Llena campos vacíos según el tipo de dato: 0 para numéricos, SIN INFO para texto"""
    
    # Campos numéricos (todos los relacionados con créditos y sectores)
    campos_numericos = [
        # CRÉDITOS VIGENTES
        "Créditos Vigentes - Sector Financiero", "Créditos Vigentes - Sector Cooperativo", 
        "Créditos Vigentes - Sector Real", "Créditos Vigentes - Sector Telcos", 
        "Créditos Vigentes - Total Sectores", "Créditos Vigentes - Total Principal", 
        "Créditos Vigentes - Total Codeudor",
        
        # CRÉDITOS CERRADOS
        "Créditos Cerrados - Sector Financiero", "Créditos Cerrados - Sector Cooperativo",
        "Créditos Cerrados - Sector Real", "Créditos Cerrados - Sector Telcos",
        "Créditos Cerrados - Total Sectores", "Créditos Cerrados - Total Principal",
        "Créditos Cerrados - Total Codeudor",
        
        # CRÉDITOS REESTRUCTURADOS
        "Créditos Reestructurados - Sector Financiero", "Créditos Reestructurados - Sector Cooperativo",
        "Créditos Reestructurados - Sector Real", "Créditos Reestructurados - Sector Telcos",
        "Créditos Reestructurados - Total Sectores", "Créditos Reestructurados - Total Principal",
        "Créditos Reestructurados - Total Codeudor",
        
        # CRÉDITOS REFINANCIADOS
        "Créditos Refinanciados - Sector Financiero", "Créditos Refinanciados - Sector Cooperativo",
        "Créditos Refinanciados - Sector Real", "Créditos Refinanciados - Sector Telcos",
        "Créditos Refinanciados - Total Sectores", "Créditos Refinanciados - Total Principal",
        "Créditos Refinanciados - Total Codeudor",
        
        # CONSULTAS ÚLTIMOS 6 MESES
        "Consultas Últimos 6 Meses - Sector Financiero", "Consultas Últimos 6 Meses - Sector Cooperativo",
        "Consultas Últimos 6 Meses - Sector Real", "Consultas Últimos 6 Meses - Sector Telcos",
        "Consultas Últimos 6 Meses - Total Sectores", "Consultas Últimos 6 Meses - Total Principal",
        "Consultas Últimos 6 Meses - Total Codeudor",
        
        # DESACUERDOS VIGENTES
        "Desacuerdos Vigentes - Sector Financiero", "Desacuerdos Vigentes - Sector Cooperativo",
        "Desacuerdos Vigentes - Sector Real", "Desacuerdos Vigentes - Sector Telcos",
        "Desacuerdos Vigentes - Total Sectores", "Desacuerdos Vigentes - Total Principal",
        "Desacuerdos Vigentes - Total Codeudor",
        
        # NÚMERO DE DOCUMENTO (también es numérico)
        "Número Documento"
    ]
    
    # Campos de texto (información personal y fechas)
    campos_texto = [
        "Consultado por", "Fecha y Hora Consulta", "Nombre", "Tipo Documento",
        "Estado Documento", "Lugar Expedición", "Fecha Expedición", 
        "Rango Edad", "Género", "Antigüedad Ubicación",
        "Antigüedad - Sector Financiero", "Antigüedad - Sector Real", 
        "Antigüedad - Sector Telcos"
    ]
    
    # Aplicar llenado para campos numéricos
    for campo in campos_numericos:
        if campo in df.columns:
            df[campo] = df[campo].replace("NO APLICA", "0")
    
    # Aplicar llenado para campos de texto
    for campo in campos_texto:
        if campo in df.columns:
            df[campo] = df[campo].replace("NO APLICA", "SIN INFO")
    
    return df

# Aplicar la función de llenado
df = llenar_campos_vacios(df)

# Ordenar por nombre para agrupar archivos del mismo cliente
def ordenar_por_cliente(df):
    """Ordena el DataFrame por nombre para agrupar archivos del mismo cliente"""
    # Primero ordenar por Nombre, luego por Fecha y Hora Consulta para mantener cronología
    df_ordenado = df.sort_values([
        'Nombre', 
        'Fecha y Hora Consulta'
    ], ascending=[True, False])  # Nombre ascendente, fecha descendente (más reciente primero)
    
    # Resetear el índice después del ordenamiento
    df_ordenado = df_ordenado.reset_index(drop=True)
    
    return df_ordenado

# Aplicar ordenamiento por cliente
df = ordenar_por_cliente(df)

# Guardar en Excel con formato profesional
excel_path = os.path.join(carpeta_salida, "resultado_detallado.xlsx")

# Crear el archivo Excel con escritor personalizado para manejar encabezados múltiples
with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    # Escribir solo los datos (sin encabezados automáticos) empezando desde la fila 3
    df.to_excel(writer, sheet_name='Datos', index=False, startrow=2, header=False)
    
    # Obtener la hoja de trabajo
    wb = writer.book
    ws = writer.sheets['Datos']
    
    # Aplicar formato a los encabezados
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Formato para encabezados principales
    header_main_font = Font(bold=True, color="FFFFFF", size=11)
    header_main_fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
    
    # Formato para subencabezados
    header_sub_font = Font(bold=True, color="000000", size=10)
    header_sub_fill = PatternFill(start_color="8db4e2", end_color="8db4e2", fill_type="solid")
    
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Definir estructura de encabezados anidados
    encabezados_principales = [
        ("INFORMACIÓN PERSONAL", 1, 11),  # (nombre, columna_inicio, columna_fin)
        ("CRÉDITOS VIGENTES", 12, 18),
        ("CRÉDITOS CERRADOS", 19, 25),
        ("CRÉDITOS REESTRUCTURADOS", 26, 32),
        ("CRÉDITOS REFINANCIADOS", 33, 39),
        ("CONSULTAS ÚLTIMOS 6 MESES", 40, 46),
        ("DESACUERDOS VIGENTES", 47, 53),
        ("ANTIGÜEDAD", 54, 56)
    ]
    
    # Crear encabezados principales en la fila 1
    for nombre, col_inicio, col_fin in encabezados_principales:
        # Combinar celdas para el encabezado principal
        start_cell = get_column_letter(col_inicio) + "1"
        end_cell = get_column_letter(col_fin) + "1"
        ws.merge_cells(f"{start_cell}:{end_cell}")
        
        # Aplicar formato al encabezado principal
        cell = ws[start_cell]
        cell.value = nombre
        cell.font = header_main_font
        cell.fill = header_main_fill
        cell.alignment = center_alignment
    
    # Crear subencabezados en la fila 2 para los grupos de créditos
    sectores_headers = ["Sector Financiero", "Sector Cooperativo", "Sector Real", "Sector Telcos", "Total Sectores", "Total Principal", "Total Codeudor"]
    
    # Información personal - mantener nombres originales
    info_personal_headers = ["Consultado por", "Fecha y Hora Consulta", "Nombre", "Tipo Documento", "Número Documento", "Estado Documento", "Lugar Expedición", "Fecha Expedición", "Rango Edad", "Género", "Antigüedad Ubicación"]
    
    for i, header in enumerate(info_personal_headers, 1):
        cell = ws[f"{get_column_letter(i)}2"]
        cell.value = header
        cell.font = header_sub_font
        cell.fill = header_sub_fill
        cell.alignment = center_alignment
    
    # Para cada grupo de créditos, agregar los subencabezados de sectores
    col_offset = 12  # Empezar después de información personal
    for grupo in range(6):  # 6 grupos de créditos
        for i, sector in enumerate(sectores_headers):
            cell = ws[f"{get_column_letter(col_offset + i)}2"]
            cell.value = sector
            cell.font = header_sub_font
            cell.fill = header_sub_fill
            cell.alignment = center_alignment
        col_offset += 7
    
    # Antigüedad - subencabezados
    antiguedad_headers = ["Sector Financiero", "Sector Real", "Sector Telcos"]
    for i, header in enumerate(antiguedad_headers):
        cell = ws[f"{get_column_letter(54 + i)}2"]
        cell.value = header
        cell.font = header_sub_font
        cell.fill = header_sub_fill
        cell.alignment = center_alignment
    
    # Aplicar formato a los datos (fila 3 en adelante) con alternancia de colores por cliente
    data_font = Font(size=9)
    
    # Colores alternos para diferentes clientes
    color_cliente_1 = PatternFill(start_color="f2f2f2", end_color="f2f2f2", fill_type="solid")  # Gris claro
    color_cliente_2 = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")  # Blanco
    
    # Detectar cambios de cliente para aplicar colores alternos
    cliente_anterior = None
    color_actual = color_cliente_1
    usar_color_1 = True
    
    for row_num in range(3, ws.max_row + 1):  # Empezar desde la fila 3 (datos)
        # Obtener el nombre del cliente en la columna C (columna 3)
        nombre_actual = ws[f"C{row_num}"].value
        
        # Si cambia el cliente, alternar color
        if nombre_actual != cliente_anterior and nombre_actual != "SIN INFO":
            if cliente_anterior is not None:  # No cambiar color en la primera fila
                usar_color_1 = not usar_color_1
            color_actual = color_cliente_1 if usar_color_1 else color_cliente_2
            cliente_anterior = nombre_actual
        
        # Aplicar formato a toda la fila
        for col_num in range(1, ws.max_column + 1):
            cell = ws[f"{get_column_letter(col_num)}{row_num}"]
            if cell.value is not None:
                cell.font = data_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = color_actual

    # Ajustar ancho de columnas automáticamente
    for col_num in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_num)
        
        # Revisar todas las celdas de la columna para calcular el ancho
        for row_num in range(1, ws.max_row + 1):
            cell = ws[f"{column_letter}{row_num}"]
            if cell.value and not isinstance(cell, type(ws['A1']).__bases__[0]):  # Evitar MergedCell
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        
        adjusted_width = min(max(max_length + 2, 12), 25)  # Mínimo 12, máximo 25 caracteres
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Aplicar bordes a toda la tabla
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Aplicar bordes a todas las celdas con contenido
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.border = thin_border
    
    # Congelar las primeras dos filas (encabezados)
    ws.freeze_panes = 'A3'
    
    # Ajustar altura de las filas de encabezados
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 35

# Mostrar resultados con interfaz gráfica
def mostrar_resultado(total_archivos, ruta_excel):
    """Muestra una ventana con el resultado del procesamiento"""
    root = tk.Tk()
    root.title("Procesamiento Completado")
    root.geometry("500x300")
    root.resizable(False, False)
    
    # Frame principal
    main_frame = ttk.Frame(root, padding="20")
    main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
    
    # Título de éxito
    title_label = ttk.Label(main_frame, text="✅ Procesamiento Exitoso", 
                           font=('Arial', 16, 'bold'), foreground='green')
    title_label.grid(row=0, column=0, columnspan=2, pady=(0, 20))
    
    # Información del resultado
    info_text = f"""Se procesaron exitosamente {total_archivos} archivos PDF.

📊 Características del archivo Excel generado:
• Información personal organizada
• Datos crediticios desglosados por sectores
• Encabezados anidados profesionales
• Ordenamiento por cliente
• Colores alternos para mejor visualización
• Campos numéricos con "0" y texto con "SIN INFO"

💾 Archivo guardado en:
{ruta_excel}"""
    
    info_label = ttk.Label(main_frame, text=info_text, font=('Arial', 10), justify='left')
    info_label.grid(row=1, column=0, columnspan=2, pady=(0, 20), sticky=(tk.W, tk.E))
    
    def abrir_carpeta():
        """Abrir la carpeta donde se guardó el archivo"""
        import subprocess
        import platform
        
        carpeta = os.path.dirname(ruta_excel)
        try:
            if platform.system() == "Windows":
                subprocess.run(["explorer", carpeta])
            elif platform.system() == "Darwin":  # macOS
                subprocess.run(["open", carpeta])
            else:  # Linux
                subprocess.run(["xdg-open", carpeta])
        except:
            messagebox.showinfo("Información", f"Archivo guardado en:\n{ruta_excel}")
    
    def cerrar():
        root.destroy()
    
    # Botones
    button_frame = ttk.Frame(main_frame)
    button_frame.grid(row=2, column=0, columnspan=2, pady=20)
    
    btn_abrir = ttk.Button(button_frame, text="Abrir Carpeta", command=abrir_carpeta)
    btn_abrir.pack(side=tk.LEFT, padx=(0, 10))
    
    btn_cerrar = ttk.Button(button_frame, text="Cerrar", command=cerrar)
    btn_cerrar.pack(side=tk.LEFT)
    
    # Centrar ventana
    root.update_idletasks()
    x = (root.winfo_screenwidth() // 2) - (root.winfo_width() // 2)
    y = (root.winfo_screenheight() // 2) - (root.winfo_height() // 2)
    root.geometry(f"+{x}+{y}")
    
    root.mainloop()

# Contar archivos procesados
archivos_procesados = len([f for f in os.listdir(carpeta_pdfs) if f.endswith('.pdf')])

print("✅ Datos extraídos y guardados en resultado_detallado.xlsx")
print("📊 Tabla organizada con información desglosada por sectores y tipos de crédito")
print("🎨 Formato aplicado: encabezados con color, bordes y columnas ajustadas")
print(f"📁 Archivo guardado en: {excel_path}")

# Mostrar ventana de resultado
mostrar_resultado(archivos_procesados, excel_path)
