# -*- coding: utf-8 -*-
"""
Procesador de Excel COMPLETO - Genera tabla única con TODA la información
"""
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelProcessorCompleto:
    """Procesador que genera Excel con TODA la información en una sola tabla"""
    
    def __init__(self):
        self.encabezados_completos = [
            # Información de consulta
            'Consultado por', 'Fecha Consulta', 'Hora Consulta', 'Tipo Consulta', 
            'Codigo Consulta', 'Entidad Consultante', 'Motivo Consulta',
            
            # Información personal
            'Tipo Documento', 'Numero Documento', 'Estado Documento', 
            'Lugar Expedicion', 'Fecha Expedicion', 'Nombre Completo',
            'Primer Nombre', 'Segundo Nombre', 'Primer Apellido', 'Segundo Apellido',
            'Rango Edad', 'Edad Exacta', 'Genero', 'Estado Civil', 'Nacionalidad',
            
            # Datos demográficos
            'Direccion Residencia', 'Ciudad Residencia', 'Departamento Residencia',
            'Antiguedad Ubicacion', 'Telefono Residencia', 'Telefono Celular',
            'Email', 'Tipo Vivienda', 'Estrato',
            
            # Actividad económica
            'Actividad Economica Principal', 'Codigo CIIU', 'Sector Economico',
            'Ingresos Reportados', 'Patrimonio Reportado', 'Empresa Trabajo',
            'Cargo', 'Tiempo Laborando', 'Tipo Contrato', 'Profesion',
            'Ocupacion', 'Nivel Educativo', 'Experiencia Laboral', 'Tipo Empleado',
            
            # Historial crediticio
            'Calificacion Crediticia', 'Categoria Riesgo', 'Experiencia Crediticia',
            'Comportamiento Pago', 'Mora Maxima', 'Saldo Total Deudas',
            'Valor Mora', 'Numero Obligaciones Activas', 'Numero Obligaciones Canceladas',
            'Cupo Credito Total', 'Cupo Utilizado', 'Cupo Disponible',
            
            # Obligaciones financieras
            'Creditos Hipotecarios', 'Creditos Vehiculo', 'Tarjetas Credito',
            'Creditos Libranza', 'MicroCreditos', 'Creditos Consumo',
            'Creditos Comerciales', 'Leasing', 'Avales', 'Cartas Credito',
            
            # Centrales de riesgo
            'DataCredito Score', 'Asobancaria Score', 'Cifin Score', 'Procredito Score',
            'Reportes Negativos', 'Reportes Positivos', 'Ultima Actualizacion',
            'Score DataCredito', 'Interpretacion Score', 'Probabilidad Incumplimiento',
            'Nivel Riesgo',
            
            # Información judicial
            'Procesos Judiciales', 'Embargos', 'Demandas', 'Remates',
            'Concordatos', 'Reestructuraciones',
            
            # Historial de consultas
            'Consultas Ultimo Mes', 'Consultas Ultimos 3 Meses', 
            'Consultas Ultimo Semestre', 'Consultas Ultimo Año',
            'Total Consultas', 'Entidades Consultantes',
            
            # Referencias
            'Referencias Bancarias', 'Referencias Comerciales', 'Referencias Personales',
            'Experiencia Sector Financiero',
            
            # Alertas y observaciones
            'Alertas Seguridad', 'Observaciones Especiales', 'Restricciones',
            'Notas Aclaratorias',
            
            # Información adicional
            'Informacion Adicional 1', 'Informacion Adicional 2', 'Informacion Adicional 3',
            'Secciones Especiales', 'Datos Complementarios',
            
            # Metadatos
            'Archivo', 'Total Campos Extraidos', 'Longitud Texto'
        ]
    
    def generar_excel(self, datos: List[Dict[str, Any]], archivo_salida: str):
        """
        Genera archivo Excel con TODA la información en una sola tabla
        
        Args:
            datos: Lista de diccionarios con datos extraídos
            archivo_salida: Ruta del archivo Excel de salida
        """
        try:
            # Crear workbook
            wb = Workbook()
            
            # Eliminar hoja por defecto
            wb.remove(wb.active)
            
            # Crear hoja única con TODA la información
            self.crear_hoja_completa(wb, datos)
            
            # Crear hoja de resumen
            self.crear_hoja_resumen(wb, datos)
            
            # Crear hoja de diagnóstico
            self.crear_hoja_diagnostico(wb, datos)
            
            # Guardar archivo
            wb.save(archivo_salida)
            print(f"\nArchivo Excel generado: {archivo_salida}")
            
        except Exception as e:
            print(f"Error generando Excel: {e}")
            raise
    
    def crear_hoja_completa(self, wb: Workbook, datos: List[Dict[str, Any]]):
        """Crea hoja con TODA la información extraída"""
        ws = wb.create_sheet("Información Completa")
        
        # Escribir encabezados
        for col, encabezado in enumerate(self.encabezados_completos, 1):
            celda = ws.cell(row=1, column=col, value=encabezado)
            self.aplicar_estilo_encabezado(celda)
        
        # Escribir datos
        for fila, registro in enumerate(datos, 2):
            for col, campo in enumerate(self.encabezados_completos, 1):
                valor = self.obtener_valor_campo(registro, campo)
                celda = ws.cell(row=fila, column=col, value=valor)
                self.aplicar_estilo_datos(celda)
        
        # Ajustar ancho de columnas
        self.ajustar_columnas(ws)
        
        # Congelar primera fila
        ws.freeze_panes = "A2"
    
    def obtener_valor_campo(self, registro: Dict[str, Any], campo: str) -> str:
        """Obtiene el valor de un campo del registro"""
        # Mapear nombres de encabezados a claves de datos
        mapeo_campos = {
            'Consultado por': 'consultado_por',
            'Fecha Consulta': 'fecha_consulta',
            'Hora Consulta': 'hora_consulta',
            'Tipo Consulta': 'tipo_consulta',
            'Codigo Consulta': 'codigo_consulta',
            'Entidad Consultante': 'entidad_consultante',
            'Motivo Consulta': 'motivo_consulta',
            
            'Tipo Documento': 'tipo_documento',
            'Numero Documento': 'numero_documento',
            'Estado Documento': 'estado_documento',
            'Lugar Expedicion': 'lugar_expedicion',
            'Fecha Expedicion': 'fecha_expedicion',
            'Nombre Completo': 'nombre_completo',
            'Primer Nombre': 'primer_nombre',
            'Segundo Nombre': 'segundo_nombre',
            'Primer Apellido': 'primer_apellido',
            'Segundo Apellido': 'segundo_apellido',
            'Rango Edad': 'rango_edad',
            'Edad Exacta': 'edad_exacta',
            'Genero': 'genero',
            'Estado Civil': 'estado_civil',
            'Nacionalidad': 'nacionalidad',
            
            'Direccion Residencia': 'direccion_residencia',
            'Ciudad Residencia': 'ciudad_residencia',
            'Departamento Residencia': 'departamento_residencia',
            'Antiguedad Ubicacion': 'antiguedad_ubicacion',
            'Telefono Residencia': 'telefono_residencia',
            'Telefono Celular': 'telefono_celular',
            'Email': 'email',
            'Tipo Vivienda': 'tipo_vivienda',
            'Estrato': 'estrato',
            
            'Actividad Economica Principal': 'actividad_economica_principal',
            'Codigo CIIU': 'codigo_ciiu',
            'Sector Economico': 'sector_economico',
            'Ingresos Reportados': 'ingresos_reportados',
            'Patrimonio Reportado': 'patrimonio_reportado',
            'Empresa Trabajo': 'empresa_trabajo',
            'Cargo': 'cargo',
            'Tiempo Laborando': 'tiempo_laborando',
            'Tipo Contrato': 'tipo_contrato',
            'Profesion': 'profesion',
            'Ocupacion': 'ocupacion',
            'Nivel Educativo': 'nivel_educativo',
            'Experiencia Laboral': 'experiencia_laboral',
            'Tipo Empleado': 'tipo_empleado',
            
            'Calificacion Crediticia': 'calificacion_crediticia',
            'Categoria Riesgo': 'categoria_riesgo',
            'Experiencia Crediticia': 'experiencia_crediticia',
            'Comportamiento Pago': 'comportamiento_pago',
            'Mora Maxima': 'mora_maxima',
            'Saldo Total Deudas': 'saldo_total_deudas',
            'Valor Mora': 'valor_mora',
            'Numero Obligaciones Activas': 'numero_obligaciones_activas',
            'Numero Obligaciones Canceladas': 'numero_obligaciones_canceladas',
            'Cupo Credito Total': 'cupo_credito_total',
            'Cupo Utilizado': 'cupo_utilizado',
            'Cupo Disponible': 'cupo_disponible',
            
            'Creditos Hipotecarios': 'creditos_hipotecarios',
            'Creditos Vehiculo': 'creditos_vehiculo',
            'Tarjetas Credito': 'tarjetas_credito',
            'Creditos Libranza': 'creditos_libranza',
            'MicroCreditos': 'microCreditos',
            'Creditos Consumo': 'creditos_consumo',
            'Creditos Comerciales': 'creditos_comerciales',
            'Leasing': 'leasing',
            'Avales': 'avales',
            'Cartas Credito': 'cartas_credito',
            
            'DataCredito Score': 'datacredito_score',
            'Asobancaria Score': 'asobancaria_score',
            'Cifin Score': 'cifin_score',
            'Procredito Score': 'procredito_score',
            'Reportes Negativos': 'reportes_negativos',
            'Reportes Positivos': 'reportes_positivos',
            'Ultima Actualizacion': 'ultima_actualizacion',
            'Score DataCredito': 'score_datacredito',
            'Interpretacion Score': 'interpretacion_score',
            'Probabilidad Incumplimiento': 'probabilidad_incumplimiento',
            'Nivel Riesgo': 'nivel_riesgo',
            
            'Procesos Judiciales': 'procesos_judiciales',
            'Embargos': 'embargos',
            'Demandas': 'demandas',
            'Remates': 'remates',
            'Concordatos': 'concordatos',
            'Reestructuraciones': 'reestructuraciones',
            
            'Consultas Ultimo Mes': 'consultas_ultimo_mes',
            'Consultas Ultimos 3 Meses': 'consultas_ultimos_3_meses',
            'Consultas Ultimo Semestre': 'consultas_ultimo_semestre',
            'Consultas Ultimo Año': 'consultas_ultimo_año',
            'Total Consultas': 'total_consultas',
            'Entidades Consultantes': 'entidades_consultantes',
            
            'Referencias Bancarias': 'referencias_bancarias',
            'Referencias Comerciales': 'referencias_comerciales',
            'Referencias Personales': 'referencias_personales',
            'Experiencia Sector Financiero': 'experiencia_sector_financiero',
            
            'Alertas Seguridad': 'alertas_seguridad',
            'Observaciones Especiales': 'observaciones_especiales',
            'Restricciones': 'restricciones',
            'Notas Aclaratorias': 'notas_aclaratorias',
            
            'Informacion Adicional 1': 'informacion_adicional_1',
            'Informacion Adicional 2': 'informacion_adicional_2',
            'Informacion Adicional 3': 'informacion_adicional_3',
            'Secciones Especiales': 'secciones_especiales',
            'Datos Complementarios': 'datos_complementarios',
            
            'Archivo': '_metadata.archivo',
            'Total Campos Extraidos': '_metadata.total_campos_extraidos',
            'Longitud Texto': '_metadata.texto_completo_length'
        }
        
        clave = mapeo_campos.get(campo, campo.lower().replace(' ', '_'))
        
        # Manejar claves anidadas (como _metadata.archivo)
        if '.' in clave:
            partes = clave.split('.')
            valor = registro
            for parte in partes:
                if isinstance(valor, dict) and parte in valor:
                    valor = valor[parte]
                else:
                    valor = ""
                    break
        else:
            valor = registro.get(clave, "")
        
        # Convertir a string y limpiar
        if valor is None:
            return ""
        return str(valor).strip()
    
    def crear_hoja_resumen(self, wb: Workbook, datos: List[Dict[str, Any]]):
        """Crea hoja de resumen estadístico"""
        ws = wb.create_sheet("Resumen Estadístico")
        
        # Título
        ws.cell(row=1, column=1, value="RESUMEN DE EXTRACCIÓN DE DATOS").font = Font(size=14, bold=True)
        
        # Estadísticas generales
        ws.cell(row=3, column=1, value="Total de archivos procesados:").font = Font(bold=True)
        ws.cell(row=3, column=2, value=len(datos))
        
        ws.cell(row=4, column=1, value="Total de campos por archivo:").font = Font(bold=True)
        ws.cell(row=4, column=2, value=len(self.encabezados_completos))
        
        ws.cell(row=5, column=1, value="Archivos con datos:").font = Font(bold=True)
        archivos_con_datos = len([d for d in datos if d.get('_metadata', {}).get('total_campos_extraidos', 0) > 0])
        ws.cell(row=5, column=2, value=archivos_con_datos)
        
        ws.cell(row=6, column=1, value="Archivos con errores:").font = Font(bold=True)
        archivos_con_errores = len([d for d in datos if not d.get('_metadata', {}).get('procesado', True)])
        ws.cell(row=6, column=2, value=archivos_con_errores)
        
        # Lista de archivos procesados
        ws.cell(row=8, column=1, value="ARCHIVOS PROCESADOS:").font = Font(bold=True)
        
        for i, registro in enumerate(datos, 9):
            metadata = registro.get('_metadata', {})
            archivo = metadata.get('archivo', 'Desconocido')
            campos = metadata.get('total_campos_extraidos', 0)
            
            ws.cell(row=i, column=1, value=archivo)
            ws.cell(row=i, column=2, value=f"{campos} campos extraídos")
            
            if not metadata.get('procesado', True):
                ws.cell(row=i, column=3, value="ERROR")
                ws.cell(row=i, column=3).font = Font(color="FF0000")
    
    def crear_hoja_diagnostico(self, wb: Workbook, datos: List[Dict[str, Any]]):
        """Crea hoja de diagnóstico técnico"""
        ws = wb.create_sheet("Diagnóstico Técnico")
        
        # Información del sistema
        ws.cell(row=1, column=1, value="DIAGNÓSTICO TÉCNICO - EXTRACTOR COMPLETO").font = Font(size=14, bold=True)
        
        ws.cell(row=3, column=1, value="Versión del extractor:").font = Font(bold=True)
        ws.cell(row=3, column=2, value="Extractor Completo v1.0")
        
        ws.cell(row=4, column=1, value="Campos configurados:").font = Font(bold=True)
        ws.cell(row=4, column=2, value=len(self.encabezados_completos))
        
        # Análisis de extracción por campo
        ws.cell(row=6, column=1, value="ANÁLISIS POR CAMPO:").font = Font(bold=True)
        
        # Encabezados
        ws.cell(row=7, column=1, value="Campo").font = Font(bold=True)
        ws.cell(row=7, column=2, value="Extracciones Exitosas").font = Font(bold=True)
        ws.cell(row=7, column=3, value="% Éxito").font = Font(bold=True)
        
        total_registros = len(datos)
        
        for i, campo in enumerate(self.encabezados_completos[:20], 8):  # Primeros 20 campos
            extracciones_exitosas = 0
            for registro in datos:
                valor = self.obtener_valor_campo(registro, campo)
                if valor and valor.strip():
                    extracciones_exitosas += 1
            
            porcentaje = (extracciones_exitosas / total_registros * 100) if total_registros > 0 else 0
            
            ws.cell(row=i, column=1, value=campo)
            ws.cell(row=i, column=2, value=extracciones_exitosas)
            ws.cell(row=i, column=3, value=f"{porcentaje:.1f}%")
    
    def aplicar_estilo_encabezado(self, celda):
        """Aplica estilo a los encabezados"""
        celda.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        celda.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        celda.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def aplicar_estilo_datos(self, celda):
        """Aplica estilo a las celdas de datos"""
        celda.font = Font(name='Calibri', size=10)
        celda.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        celda.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def ajustar_columnas(self, ws):
        """Ajusta el ancho de las columnas"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Limitar ancho mínimo y máximo
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width