# -*- coding: utf-8 -*-
"""
Procesador para generar archivos Excel con datos extraídos
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from typing import List, Dict, Any

class ExcelProcessor:
    """Procesador para generar archivos Excel formateados"""
    
    def __init__(self):
        # Estilos para Excel
        self.estilo_encabezado = {
            'font': Font(bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
        
        self.estilo_celda = {
            'alignment': Alignment(horizontal='left', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    def generar_excel(self, datos: List[Dict[str, Any]], archivo_salida: str):
        """
        Genera archivo Excel con múltiples hojas
        
        Args:
            datos: Lista de diccionarios con datos extraídos
            archivo_salida: Ruta del archivo Excel de salida
        """
        try:
            # Crear workbook
            wb = Workbook()
            
            # Eliminar hoja por defecto
            wb.remove(wb.active)
            
            # Generar hoja principal para el usuario
            self.crear_hoja_informacion_basica(wb, datos)
            
            # Generar hoja de resumen
            self.crear_hoja_resumen(wb, datos)
            
            # Generar hoja de diagnóstico técnico
            self.crear_hoja_diagnostico(wb, datos)
            
            # Guardar archivo
            wb.save(archivo_salida)
            print(f"\nArchivo Excel generado: {archivo_salida}")
            
        except Exception as e:
            print(f"Error generando Excel: {e}")
            raise
    
    def crear_hoja_informacion_basica(self, wb: Workbook, datos: List[Dict[str, Any]]):
        """Crea hoja con información básica"""
        ws = wb.create_sheet("Información Básica")
        
        # Encabezados (sin mostrar columna de Archivo al usuario final)
        encabezados = [
            "Consultado por", "Fecha y Hora Consulta", "Tipo Documento",
            "Número Documento", "Estado Documento", "Lugar Expedición",
            "Fecha Expedición", "Nombre", "Rango Edad", "Género",
            "Antigüedad Ubicación"
        ]
        
        # Escribir encabezados
        for col, encabezado in enumerate(encabezados, 1):
            celda = ws.cell(row=1, column=col, value=encabezado)
            self.aplicar_estilo_encabezado(celda)
        
        # Escribir datos (sin columna de archivo, procesado o error para usuario final)
        for fila, registro in enumerate(datos, 2):
            info_basica = registro.get('informacion_basica', {})
            
            valores = [
                info_basica.get('Consultado por', ''),
                info_basica.get('Fecha y Hora Consulta', ''),
                info_basica.get('Tipo Documento', ''),
                info_basica.get('Número Documento', ''),
                info_basica.get('Estado Documento', ''),
                info_basica.get('Lugar Expedición', ''),
                info_basica.get('Fecha Expedición', ''),
                info_basica.get('Nombre', ''),
                info_basica.get('Rango Edad', ''),
                info_basica.get('Género', ''),
                info_basica.get('Antigüedad Ubicación', '')
            ]
            
            for col, valor in enumerate(valores, 1):
                celda = ws.cell(row=fila, column=col, value=valor)
                self.aplicar_estilo_celda(celda)
        
        # Ajustar anchos de columna
        self.ajustar_anchos_columna(ws)
    
    def crear_hoja_resumen(self, wb: Workbook, datos: List[Dict[str, Any]]):
        """Crea hoja con resumen estadístico"""
        ws = wb.create_sheet("Resumen")
        
        # Calcular estadísticas
        total = len(datos)
        procesados = sum(1 for d in datos if d.get('_metadata', {}).get('procesado', False))
        errores = total - procesados
        
        # Datos del resumen
        resumen_datos = [
            ["Estadística", "Valor"],
            ["Total de archivos", total],
            ["Procesados exitosamente", procesados],
            ["Con errores", errores],
            ["Porcentaje de éxito", f"{(procesados/total*100):.1f}%" if total > 0 else "0%"]
        ]
        
        # Escribir resumen
        for fila, (etiqueta, valor) in enumerate(resumen_datos, 1):
            ws.cell(row=fila, column=1, value=etiqueta)
            ws.cell(row=fila, column=2, value=valor)
        
        # Formatear encabezados
        for col in range(1, 3):
            celda = ws.cell(row=1, column=col)
            self.aplicar_estilo_encabezado(celda)
        
        # Formatear datos
        for fila in range(2, len(resumen_datos) + 1):
            for col in range(1, 3):
                celda = ws.cell(row=fila, column=col)
                self.aplicar_estilo_celda(celda)
        
        # Lista de archivos con errores
        if errores > 0:
            ws.cell(row=len(resumen_datos) + 3, column=1, value="Archivos con errores:")
            fila_error = len(resumen_datos) + 4
            
            for registro in datos:
                metadata = registro.get('_metadata', {})
                if not metadata.get('procesado', False):
                    ws.cell(row=fila_error, column=1, value=metadata.get('archivo', ''))
                    ws.cell(row=fila_error, column=2, value=metadata.get('error', ''))
                    fila_error += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
    
    def crear_hoja_diagnostico(self, wb: Workbook, datos: List[Dict[str, Any]]):
        """Crea hoja con información técnica de diagnóstico"""
        ws = wb.create_sheet("Diagnóstico Técnico")
        
        # Encabezados técnicos
        encabezados = [
            "Archivo", "Consultado por", "Fecha y Hora Consulta", "Tipo Documento",
            "Número Documento", "Estado Documento", "Lugar Expedición",
            "Fecha Expedición", "Nombre", "Rango Edad", "Género",
            "Antigüedad Ubicación", "Procesado", "Error"
        ]
        
        # Escribir encabezados
        for col, encabezado in enumerate(encabezados, 1):
            celda = ws.cell(row=1, column=col, value=encabezado)
            self.aplicar_estilo_encabezado(celda)
        
        # Escribir datos técnicos completos
        for fila, registro in enumerate(datos, 2):
            info_basica = registro.get('informacion_basica', {})
            metadata = registro.get('_metadata', {})
            
            valores = [
                metadata.get('archivo', ''),
                info_basica.get('Consultado por', ''),
                info_basica.get('Fecha y Hora Consulta', ''),
                info_basica.get('Tipo Documento', ''),
                info_basica.get('Número Documento', ''),
                info_basica.get('Estado Documento', ''),
                info_basica.get('Lugar Expedición', ''),
                info_basica.get('Fecha Expedición', ''),
                info_basica.get('Nombre', ''),
                info_basica.get('Rango Edad', ''),
                info_basica.get('Género', ''),
                info_basica.get('Antigüedad Ubicación', ''),
                'Sí' if metadata.get('procesado', False) else 'No',
                metadata.get('error', '')
            ]
            
            for col, valor in enumerate(valores, 1):
                celda = ws.cell(row=fila, column=col, value=valor)
                self.aplicar_estilo_celda(celda)
        
        # Ajustar anchos de columna
        self.ajustar_anchos_columna(ws)
    
    def aplicar_estilo_encabezado(self, celda):
        """Aplica estilo a celda de encabezado"""
        celda.font = self.estilo_encabezado['font']
        celda.fill = self.estilo_encabezado['fill']
        celda.alignment = self.estilo_encabezado['alignment']
        celda.border = self.estilo_encabezado['border']
    
    def aplicar_estilo_celda(self, celda):
        """Aplica estilo a celda normal"""
        celda.alignment = self.estilo_celda['alignment']
        celda.border = self.estilo_celda['border']
    
    def ajustar_anchos_columna(self, ws):
        """Ajusta automáticamente el ancho de las columnas"""
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
            ws.column_dimensions[column].width = adjusted_width