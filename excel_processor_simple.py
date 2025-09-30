# -*- coding: utf-8 -*-
"""
Procesador de Excel Simplificado - Sin dependencias complejas
"""
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

class ExcelProcessorSimple:
    """Procesador Excel simple y robusto"""
    
    def __init__(self):
        self.nombre = "Excel Processor Simple"
        
    def generar_excel(self, datos: List[Dict[str, Any]], archivo_salida: str):
        """Genera Excel con toda la información extraída"""
        
        try:
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Datos Extraídos"
            
            print(f"📊 Generando Excel con {len(datos)} registros...")
            
            if not datos:
                ws.cell(row=1, column=1, value="No hay datos para procesar")
                wb.save(archivo_salida)
                print(f"⚠️ Excel vacío generado: {archivo_salida}")
                return
            
            # Recopilar todos los campos únicos de todos los registros
            todos_los_campos = set()
            for registro in datos:
                for campo in registro.keys():
                    if campo != '_metadata':
                        todos_los_campos.add(campo)
            
            # Convertir a lista ordenada
            campos_ordenados = sorted(list(todos_los_campos))
            
            # Agregar campos de metadata al final
            campos_ordenados.extend(['archivo', 'total_campos_extraidos', 'procesado'])
            
            print(f"📋 Campos detectados: {len(campos_ordenados)}")
            for i, campo in enumerate(campos_ordenados[:10], 1):  # Mostrar primeros 10
                print(f"   {i:2d}. {campo}")
            if len(campos_ordenados) > 10:
                print(f"   ... y {len(campos_ordenados) - 10} campos más")
            
            # Escribir encabezados
            for col, campo in enumerate(campos_ordenados, 1):
                celda = ws.cell(row=1, column=col, value=self.limpiar_nombre_campo(campo))
                # Estilo de encabezado
                celda.font = Font(bold=True, color='FFFFFF')
                celda.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                celda.alignment = Alignment(horizontal='center', wrap_text=True)
            
            # Escribir datos
            for fila, registro in enumerate(datos, 2):
                for col, campo in enumerate(campos_ordenados, 1):
                    valor = self.obtener_valor_campo(registro, campo)
                    ws.cell(row=fila, column=col, value=valor)
            
            # Ajustar ancho de columnas (básico)
            for col in range(1, len(campos_ordenados) + 1):
                ws.column_dimensions[self.get_column_letter(col)].width = 15
            
            # Congelar primera fila
            ws.freeze_panes = "A2"
            
            # Crear hoja de resumen
            self.crear_hoja_resumen(wb, datos)
            
            # Guardar
            wb.save(archivo_salida)
            print(f"✅ Excel generado exitosamente: {archivo_salida}")
            print(f"   • {len(datos)} registros procesados")
            print(f"   • {len(campos_ordenados)} campos por registro")
            
        except Exception as e:
            print(f"❌ Error generando Excel: {e}")
            # Crear Excel básico de emergencia
            try:
                wb = Workbook()
                ws = wb.active
                ws.cell(row=1, column=1, value="Error en generación")
                ws.cell(row=2, column=1, value=str(e))
                wb.save(archivo_salida)
                print(f"📄 Excel de error generado: {archivo_salida}")
            except:
                pass
            raise
    
    def obtener_valor_campo(self, registro: Dict[str, Any], campo: str) -> str:
        """Obtiene el valor de un campo del registro"""
        
        # Campos de metadata especiales
        if campo == 'archivo':
            return registro.get('_metadata', {}).get('archivo', '')
        elif campo == 'total_campos_extraidos':
            return str(registro.get('_metadata', {}).get('total_campos_extraidos', 0))
        elif campo == 'procesado':
            return 'Sí' if registro.get('_metadata', {}).get('procesado', False) else 'No'
        
        # Campo normal
        valor = registro.get(campo, '')
        
        if valor is None:
            return ''
        
        # Limpiar y truncar valor si es muy largo
        valor_str = str(valor).strip()
        if len(valor_str) > 200:  # Truncar valores muy largos
            valor_str = valor_str[:197] + "..."
        
        return valor_str
    
    def limpiar_nombre_campo(self, campo: str) -> str:
        """Limpia el nombre del campo para ser más legible"""
        # Reemplazar guiones bajos con espacios y capitalizar
        campo_limpio = campo.replace('_', ' ').title()
        
        # Algunos nombres específicos
        reemplazos = {
            'Consultado Por': 'Consultado por',
            'Fecha Consulta': 'Fecha Consulta',
            'Numero Documento': 'Número Documento',
            'Direccion Residencia': 'Dirección Residencia',
            'Auto ': '',  # Remover prefijo "Auto"
        }
        
        for original, reemplazo in reemplazos.items():
            campo_limpio = campo_limpio.replace(original, reemplazo)
        
        return campo_limpio
    
    def crear_hoja_resumen(self, wb: Workbook, datos: List[Dict[str, Any]]):
        """Crea hoja de resumen"""
        ws = wb.create_sheet("Resumen")
        
        # Título
        ws.cell(row=1, column=1, value="RESUMEN DE EXTRACCIÓN").font = Font(size=14, bold=True)
        
        # Estadísticas
        ws.cell(row=3, column=1, value="Total archivos procesados:").font = Font(bold=True)
        ws.cell(row=3, column=2, value=len(datos))
        
        exitosos = len([d for d in datos if d.get('_metadata', {}).get('procesado', False)])
        ws.cell(row=4, column=1, value="Archivos con datos extraídos:").font = Font(bold=True)
        ws.cell(row=4, column=2, value=exitosos)
        
        errores = len(datos) - exitosos
        ws.cell(row=5, column=1, value="Archivos con errores:").font = Font(bold=True)
        ws.cell(row=5, column=2, value=errores)
        
        # Lista de archivos
        ws.cell(row=7, column=1, value="ARCHIVOS PROCESADOS:").font = Font(bold=True)
        for i, registro in enumerate(datos, 8):
            metadata = registro.get('_metadata', {})
            archivo = metadata.get('archivo', f'Archivo {i-7}')
            campos = metadata.get('total_campos_extraidos', 0)
            
            ws.cell(row=i, column=1, value=archivo)
            ws.cell(row=i, column=2, value=f"{campos} campos")
            
            if not metadata.get('procesado', True):
                ws.cell(row=i, column=3, value="ERROR").font = Font(color='FF0000')
    
    def get_column_letter(self, col_idx):
        """Convierte número de columna a letra"""
        result = ""
        while col_idx > 0:
            col_idx -= 1
            result = chr(col_idx % 26 + ord('A')) + result
            col_idx = col_idx // 26
        return result