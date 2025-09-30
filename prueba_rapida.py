# -*- coding: utf-8 -*-
"""
PRUEBA RÁPIDA - Verificar extracción con texto de ejemplo
"""
import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from extractor_independiente import ExtractorIndependiente

def probar_con_texto_ejemplo():
    """Prueba el extractor con texto de ejemplo típico de DataCrédito"""
    
    # Texto de ejemplo simulando un PDF DataCrédito
    texto_ejemplo = """
    DATACREDITO - REPORTE DE CREDITO
    
    Consultado por: YURANY VARON OVIEDO DELAGRO SAS
    Fecha y Hora Consulta: 15/09/2024 14:30:25
    
    INFORMACIÓN PERSONAL
    Tipo Documento: Cédula de Ciudadanía
    Número Documento: 52.123.456
    Estado Documento: Vigente
    Lugar Expedición: IBAGUE
    Fecha Expedición: 12/03/1990
    Nombre: MARIA ELENA RODRIGUEZ GARCIA
    Rango Edad: 30-35
    Género: Femenino
    Antigüedad Ubicación: 24 Meses En Esta Ubicación
    
    INFORMACIÓN DE CONTACTO
    Dirección: CALLE 15 # 20-30 APTO 201
    Ciudad: IBAGUE
    Teléfono: 318-555-1234
    Email: maria.rodriguez@email.com
    
    INFORMACIÓN CREDITICIA
    Score DataCrédito: 650
    Calificación: A
    Saldo Total: $2.500.000
    Obligaciones Activas: 3
    """
    
    print("🧪 PRUEBA CON TEXTO DE EJEMPLO")
    print("=" * 50)
    
    # Crear extractor
    extractor = ExtractorIndependiente()
    
    # Extraer información
    resultado = extractor.extract(texto_ejemplo, "ejemplo.pdf")
    
    print("📊 RESULTADOS DE EXTRACCIÓN:")
    print("-" * 30)
    
    campos_encontrados = 0
    for campo, valor in resultado.items():
        if campo != '_metadata' and valor and str(valor).strip():
            print(f"✅ {campo}: {valor}")
            campos_encontrados += 1
    
    print(f"\n📈 RESUMEN:")
    print(f"   Total campos extraídos: {campos_encontrados}")
    print(f"   Metadatos: {resultado.get('_metadata', {})}")
    
    if campos_encontrados > 0:
        print("\n✅ EL EXTRACTOR FUNCIONA CORRECTAMENTE")
        print("📊 Generando Excel básico...")
        
        # Excel básico usando openpyxl directamente
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Datos Extraídos"
            
            # Encabezados
            ws.cell(row=1, column=1, value="Campo")
            ws.cell(row=1, column=2, value="Valor")
            
            # Datos
            fila = 2
            for campo, valor in resultado.items():
                if campo != '_metadata' and valor:
                    ws.cell(row=fila, column=1, value=campo)
                    ws.cell(row=fila, column=2, value=str(valor))
                    fila += 1
            
            wb.save("prueba_extraccion.xlsx")
            print("✅ Excel generado correctamente: prueba_extraccion.xlsx")
        except Exception as e:
            print(f"❌ Error generando Excel: {e}")
    else:
        print("\n❌ EL EXTRACTOR NO ENCONTRÓ INFORMACIÓN")
        print("💡 Revisando patrones...")

def main():
    probar_con_texto_ejemplo()

if __name__ == "__main__":
    main()